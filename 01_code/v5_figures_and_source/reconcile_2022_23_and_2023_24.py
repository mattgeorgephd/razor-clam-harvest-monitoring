import openpyxl, datetime, csv, collections
base="repo/02_data/harvest_inseason_records/"
BEACH={5:"Longbeach",7:"Twin Harbors",9:"Copalis",11:"Mocrocks",13:"Kalaloch"}
def parse(f):
    wbv=openpyxl.load_workbook(base+f, data_only=True)
    wbf=openpyxl.load_workbook(base+f, data_only=False)
    wv,wf=wbv["Effort"],wbf["Effort"]
    out=[]
    for r in range(2, wv.max_row+1):
        d=wv.cell(r,1).value
        if not isinstance(d, datetime.datetime): continue
        rec={"date":d.date()}
        for c,b in BEACH.items():
            rec[b]=(wv.cell(r,c).value, wv.cell(r,c+1).value, wf.cell(r,c+1).value)
        out.append(rec)
    wbv.close(); wbf.close(); return out

def dbload(season):
    db={}; 
    with open("repo/02_data/RazorClam_HarvestDB_Extract_and_Pipeline/derived/harvest_by_beach_day.csv") as f:
        for r in csv.DictReader(f):
            if r["season"]!=season: continue
            b=r["beach"].replace("LongBeach","Longbeach")
            db[(b, datetime.date.fromisoformat(r["date"][:10]))]=float(r["effort"])
    return db

for f,season in [("2022 Fall - 2023 Spring Harvest.xlsx","2022-23"),("2023 Fall - 2024 Spring Harvest.xlsx","2023-24")]:
    rows=parse(f); db=dbload(season)
    print("="*95); print(f, "|", season, "| dated rows", len(rows), "| first", rows[0]["date"], "last", rows[-1]["date"])
    tab_open=collections.Counter(); tab_eff=collections.defaultdict(float)
    matched=collections.Counter(); dbeff_matched=collections.defaultdict(float)
    est=collections.Counter()
    for r in rows:
        for b in list(BEACH.values())[:4]:
            fl,v,fr = r[b]
            if isinstance(v,(int,float)) and v>0:
                tab_open[b]+=1; tab_eff[b]+=v
                if (b,r["date"]) in db:
                    matched[b]+=1; dbeff_matched[b]+=db[(b,r["date"])]
                if fl in ("*","**"): est[b]+=1
    dbdays=collections.Counter(); dbeff=collections.defaultdict(float)
    for (b,d),e in db.items():
        dbdays[b]+=1; dbeff[b]+=e
    print(f"{'beach':14s} {'tab open bd':>11s} {'tab effort':>11s} {'flagged est':>11s} {'DB bd':>6s} {'DB eff':>10s} {'matched bd':>10s} {'DB eff on matched':>18s} {'share':>7s}")
    for b in list(BEACH.values())[:4]:
        sh = dbeff_matched[b]/tab_eff[b] if tab_eff[b] else 0
        print(f"{b:14s} {tab_open[b]:>11} {tab_eff[b]:>11,.0f} {est[b]:>11} {dbdays[b]:>6} {dbeff[b]:>10,.0f} {matched[b]:>10} {dbeff_matched[b]:>18,.0f} {sh:>7.3f}")
    # DB beach-days not open in tab
    tabset={(b,r["date"]) for r in rows for b in list(BEACH.values())[:4] if isinstance(r[b][1],(int,float)) and r[b][1]>0}
    extra=sorted([(b,d,round(e)) for (b,d),e in db.items() if (b,d) not in tabset], key=lambda x:(x[1],x[0]))
    print(f"\n  DB beach-days NOT open in the Effort tab: {len(extra)}, total {sum(x[2] for x in extra):,} digger trips")
    for b,d,e in extra: print("   ", d, b, f"{e:,}")
    cw_tab=sum(tab_eff.values()); cw_db=sum(dbeff_matched.values())
    print(f"  coastwide tab effort {cw_tab:,.0f}; DB-backed on matched days {cw_db:,.0f}; share {cw_db/cw_tab:.3f}")
